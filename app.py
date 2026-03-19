import pymysql
import pymysql.cursors

import os, json, tempfile, smtplib, threading
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
TEMP_DIR = tempfile.gettempdir()
from datetime import datetime, date, timedelta
from functools import wraps

from flask import (Flask, render_template, request, redirect,
                   url_for, session, flash, jsonify, send_file, g)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

from config import Config

try:
    import openpyxl
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

app = Flask(__name__)
app.config.from_object(Config)

# ── Pure PyMySQL connection (replaces Flask-MySQLdb) ──
def get_db():
    if 'db' not in g:
        g.db = pymysql.connect(
            host=app.config['MYSQL_HOST'],
            port=int(app.config.get('MYSQL_PORT', 3306)),
            user=app.config['MYSQL_USER'],
            password=app.config['MYSQL_PASSWORD'],
            database=app.config['MYSQL_DB'],
            cursorclass=pymysql.cursors.DictCursor,
            autocommit=False,
            connect_timeout=10,
        )
    return g.db

@app.teardown_appcontext
def close_db(error=None):
    db = g.pop('db', None)
    if db is not None:
        db.close()

class _MySQLCompat:
    """Thin shim so existing code using mysql.connection.cursor() keeps working."""
    @property
    def connection(self):
        return get_db()

mysql = _MySQLCompat()

# Jinja filter for JSON parsing in templates
app.jinja_env.filters['from_json'] = json.loads

ALLOWED_IMG  = {'jpg','jpeg','png'}
ALLOWED_PDF  = {'pdf'}
ALLOWED_CERT = {'pdf','jpg','jpeg','png'}

ALL_CLASSES = [
    'I-A','I-B','I-C','I-D','I-E','I-F','I-G',
    'CS-A','CS-B','CE-A','MECH-A','AD-A','AD-B','ECE-A'
]

# Classes grouped by year (1st year = sem 1&2, 2nd = sem 3&4, 3rd = sem 5&6, 4th = sem 7&8)
CLASSES_BY_YEAR = {
    1: ['I-A','I-B','I-C','I-D','I-E','I-F','I-G'],
    2: ['CS-A','CS-B','CE-A','MECH-A','AD-A','AD-B','ECE-A'],
    3: ['CS-A','CS-B','CE-A','MECH-A','AD-A','AD-B','ECE-A'],
    4: ['CS-A','CS-B','CE-A','MECH-A','AD-A','AD-B','ECE-A'],
}

# ════════════════════════════════════════════
#  HELPERS
# ════════════════════════════════════════════

def get_cur():
    return mysql.connection.cursor()

def parse_class_key(class_key):
    """Parse '3-AD-B' → (cls_name='AD-B', cls_year=3).
       Plain keys like 'I-A' or 'AD-B' → (cls_name=key, cls_year=None)."""
    import re as _re
    m = _re.match(r'^(\d+)-(.+)$', class_key or '')
    if m:
        return m.group(2), int(m.group(1))
    return class_key, None

def year_filter_sql(cls_year, alias=''):
    """Return (sql_fragment, params_list) for year filtering."""
    col = f"{alias}.year" if alias else "year"
    if cls_year:
        return f" AND {col}=%s", [cls_year]
    return "", []

def allowed(filename, exts):
    return '.' in filename and filename.rsplit('.',1)[1].lower() in exts

def save_file(file, subfolder):
    fn   = secure_filename(file.filename)
    ts   = datetime.now().strftime('%Y%m%d%H%M%S')
    fn   = f"{ts}_{fn}"
    path = os.path.join(app.config['UPLOAD_FOLDER'], subfolder, fn)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    file.save(path)
    return f"uploads/{subfolder}/{fn}"

def get_current_user():
    if 'user_id' not in session:
        return None
    try:
        cur = get_cur()
        cur.execute("SELECT * FROM users WHERE id=%s", (session['user_id'],))
        u = cur.fetchone(); cur.close()
        return u
    except:
        return None

def get_staff_classes(staff_id):
    """Return list of class_names this staff chose in their profile."""
    try:
        cur = get_cur()
        cur.execute("SELECT class_name FROM staff_classes WHERE staff_id=%s ORDER BY class_name",
                    (staff_id,))
        rows = cur.fetchall(); cur.close()
        return [r['class_name'] for r in rows]
    except:
        return []

def get_class_coordinator(class_name):
    """Return the staff user who is CC for this class, or None."""
    try:
        cur = get_cur()
        cur.execute("""SELECT u.* FROM users u
                       JOIN staff_classes sc ON sc.staff_id=u.id
                       WHERE sc.is_coordinator=1 AND sc.coordinator_class=%s
                       LIMIT 1""", (class_name,))
        cc = cur.fetchone(); cur.close()
        return cc
    except:
        return None

def add_stars(user_id, amount, reason, notif_type='streak'):
    try:
        cur = get_cur()
        # Don't give stars to admin
        cur.execute("SELECT role FROM users WHERE id=%s", (user_id,))
        u = cur.fetchone()
        if u and u['role'] == 'admin':
            cur.close(); return

        if amount >= 0:
            cur.execute("""UPDATE users SET streak_stars=streak_stars+%s,
                           total_stars_earned=total_stars_earned+%s WHERE id=%s""",
                        (amount, amount, user_id))
            cur.execute("INSERT INTO streak_log (user_id,stars_earned,reason) VALUES (%s,%s,%s)",
                        (user_id, amount, reason))
        else:
            cur.execute("UPDATE users SET streak_stars=GREATEST(streak_stars+%s,0) WHERE id=%s",
                        (amount, user_id))
            cur.execute("INSERT INTO streak_log (user_id,stars_deducted,reason) VALUES (%s,%s,%s)",
                        (user_id, abs(amount), reason))

        cur.execute("SELECT streak_stars FROM users WHERE id=%s", (user_id,))
        row = cur.fetchone()
        if row:
            new_level = max(1, (row['streak_stars'] or 0) // 1000 + 1)
            cur.execute("UPDATE users SET level=%s WHERE id=%s", (new_level, user_id))

        _notif(cur, user_id,
               f'+{amount} ⭐ Stars Earned' if amount >= 0 else f'{amount} ⭐ Stars Deducted',
               reason, notif_type if amount >= 0 else 'warning')
        mysql.connection.commit(); cur.close()
    except Exception as e:
        print(f"add_stars error: {e}")

def _notif(cur, user_id, title, message, ntype='general'):
    cur.execute("INSERT INTO notifications (user_id,title,message,type) VALUES (%s,%s,%s,%s)",
                (user_id, title, message, ntype))

def notify(user_id, title, message, ntype='general'):
    try:
        cur = get_cur()
        _notif(cur, user_id, title, message, ntype)
        mysql.connection.commit(); cur.close()
    except Exception as e:
        print(f"notify error: {e}")

def get_notif_count(user_id):
    try:
        cur = get_cur()
        cur.execute("SELECT COUNT(*) AS c FROM notifications WHERE user_id=%s AND is_read=0", (user_id,))
        r = cur.fetchone(); cur.close()
        return r['c'] if r else 0
    except:
        return 0

def calc_grade(pct):
    if pct >= 90: return 'O'
    if pct >= 80: return 'A+'
    if pct >= 70: return 'A'
    if pct >= 60: return 'B+'
    if pct >= 50: return 'B'
    if pct >= 40: return 'C'
    return 'F'

def get_today_slots(staff_id, class_name=None):
    day_name = date.today().strftime('%A')
    try:
        cur = get_cur()
        if class_name:
            cur.execute("""SELECT * FROM timetable WHERE staff_id=%s AND day=%s AND class_name=%s
                           ORDER BY period""", (staff_id, day_name, class_name))
        else:
            cur.execute("SELECT * FROM timetable WHERE staff_id=%s AND day=%s ORDER BY period",
                        (staff_id, day_name))
        rows = cur.fetchall(); cur.close()
        return rows
    except:
        return []

# ─── Decorators ────────────────────────────
def login_required(f):
    @wraps(f)
    def deco(*a,**kw):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*a,**kw)
    return deco

def staff_required(f):
    @wraps(f)
    def deco(*a,**kw):
        if session.get('role') not in ('staff','admin'):
            flash('Staff access required','danger')
            return redirect(url_for('dashboard'))
        return f(*a,**kw)
    return deco

def admin_required(f):
    @wraps(f)
    def deco(*a,**kw):
        if session.get('role') != 'admin':
            flash('Admin access required','danger')
            return redirect(url_for('dashboard'))
        return f(*a,**kw)
    return deco

def student_required(f):
    @wraps(f)
    def deco(*a,**kw):
        if session.get('role') != 'student':
            flash('Student access required','danger')
            return redirect(url_for('dashboard'))
        return f(*a,**kw)
    return deco

@app.context_processor
def inject_globals():
    nc = 0
    if 'user_id' in session:
        nc = get_notif_count(session['user_id'])
    return dict(notif_count=nc, today=date.today(), all_classes=ALL_CLASSES, classes_by_year=CLASSES_BY_YEAR)

# ════════════════════════════════════════════
#  AUTH
# ════════════════════════════════════════════

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email','').strip().lower()
        pwd   = request.form.get('password','')
        cur = get_cur()
        cur.execute("SELECT * FROM users WHERE email=%s", (email,))
        u = cur.fetchone(); cur.close()
        if u and check_password_hash(u['password_hash'], pwd):
            session.clear()
            session['user_id'] = u['id']
            session['role']    = u['role']
            session['name']    = u['name']
            return redirect(url_for('dashboard'))
        flash('Invalid email or password','danger')
    return render_template('auth/login.html')

@app.route('/register', methods=['GET','POST'])
def register():
    if request.method == 'POST':
        name  = request.form.get('name','').strip()
        email = request.form.get('email','').strip().lower()
        pwd   = request.form.get('password','')
        role  = request.form.get('role','student')
        dept  = request.form.get('department','')
        phone = request.form.get('phone','')
        sid   = request.form.get('staff_id','')
        is_cc      = 1 if request.form.get('is_coordinator') else 0
        cc_class   = request.form.get('coordinator_class','') if is_cc else ''
        cc_year    = request.form.get('coordinator_year', '') if is_cc else ''

        # Students get class/year/roll — staff get none of these
        if role == 'student':
            year = request.form.get('year', type=int, default=1)
            cls  = request.form.get('class_name','')
            roll = request.form.get('roll_number','').strip()
        else:
            year = None
            cls  = None
            roll = None

        if not name or not email or not pwd:
            flash('Please fill in all required fields','danger')
            return render_template('auth/register.html', all_classes=ALL_CLASSES)

        cur = get_cur()
        cur.execute("SELECT id FROM users WHERE email=%s", (email,))
        if cur.fetchone():
            flash('Email already registered — please log in instead.','danger'); cur.close()
            return render_template('auth/register.html', all_classes=ALL_CLASSES)

        ph = generate_password_hash(pwd)
        try:
            cur.execute("""INSERT INTO users (name,email,password_hash,role,department,phone,year,class_name,roll_number,staff_id)
                           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                        (name,email,ph,role,dept,phone,year,cls,roll,sid))
        except Exception as e:
            cur.close()
            if '1062' in str(e) or 'Duplicate' in str(e):
                flash('Email already registered — please log in instead.','danger')
            else:
                flash(f'Registration error: {e}','danger')
            return render_template('auth/register.html', all_classes=ALL_CLASSES)

        new_id = cur.lastrowid

        # If staff and is CC, record with year info
        if role == 'staff' and is_cc and cc_class:
            # Always store year prefix in coordinator_class e.g. "1-AD-B", "2-AD-B"
            cc_key = f"{cc_year}-{cc_class}" if cc_year else cc_class
            cur.execute("""INSERT INTO staff_classes (staff_id, class_name, is_coordinator, coordinator_class)
                           VALUES (%s,%s,1,%s)
                           ON DUPLICATE KEY UPDATE is_coordinator=1, coordinator_class=%s""",
                        (new_id, cc_key, cc_key, cc_key))

        mysql.connection.commit(); cur.close()
        flash('Account created! Please log in.','success')
        return redirect(url_for('login'))
    return render_template('auth/register.html', all_classes=ALL_CLASSES)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ════════════════════════════════════════════
#  DASHBOARD — role router
# ════════════════════════════════════════════

@app.route('/dashboard')
@login_required
def dashboard():
    user = get_current_user()
    if not user:
        session.clear()
        flash('Session expired — please log in again.','warning')
        return redirect(url_for('login'))
    role = session.get('role')
    if role == 'admin':
        return redirect(url_for('admin_dashboard'))
    elif role == 'staff':
        return redirect(url_for('staff_dashboard'))
    else:
        return _student_dashboard(user)

# ════════════════════════════════════════════
#  ADMIN PANEL
# ════════════════════════════════════════════

@app.route('/admin')
@login_required
@admin_required
def admin_dashboard():
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()

    # ── stats ──
    cur.execute("SELECT COUNT(*) AS c FROM users WHERE role='student'")
    total_students = cur.fetchone()['c']
    cur.execute("SELECT COUNT(*) AS c FROM users WHERE role='staff'")
    total_staff = cur.fetchone()['c']
    cur.execute("SELECT COUNT(*) AS c FROM attendance WHERE date=CURDATE()")
    today_att = cur.fetchone()['c']
    cur.execute("SELECT COUNT(*) AS c FROM certificates WHERE verification_status='pending'")
    pending_certs = cur.fetchone()['c']

    # ── performance per class (split by year for 2nd–4th year classes) ──
    class_stats = []
    for yr, cls_list in CLASSES_BY_YEAR.items():
        for cls in cls_list:
            year_filter = " AND u.year=%s" if yr > 1 else ""
            params = [cls, cls] if yr == 1 else [cls, yr, cls, yr]
            cur.execute("SELECT COUNT(*) AS total FROM users WHERE class_name=%s AND role='student'" +
                        ("" if yr == 1 else " AND year=%s"), ([cls] if yr == 1 else [cls, yr]))
            row = cur.fetchone()
            if not row or row['total'] == 0: continue
            total = row['total']
            cur.execute("""SELECT AVG(streak_stars) AS avg_stars,
                           AVG(CASE WHEN total_att>0 THEN present_count/total_att*100 ELSE 0 END) AS avg_att
                           FROM (
                               SELECT u.streak_stars,
                               (SELECT COUNT(*) FROM attendance WHERE student_id=u.id AND status='present') AS present_count,
                               (SELECT COUNT(*) FROM attendance WHERE student_id=u.id) AS total_att
                               FROM users u WHERE u.class_name=%s AND u.role='student'""" +
                           year_filter + ") t",
                           ([cls] if yr == 1 else [cls, yr]))
            sr = cur.fetchone()
            label = cls if yr == 1 else f"Y{yr}-{cls}"
            class_stats.append({
                'class_name': label,
                'raw_class': cls,
                'year': yr,
                'total': total,
                'avg_stars': round(sr['avg_stars'] or 0, 1),
                'avg_att': round(sr['avg_att'] or 0, 1)
            })

    # ── all staff list ──
    cur.execute("""SELECT u.*, GROUP_CONCAT(sc.class_name ORDER BY sc.class_name SEPARATOR ', ') AS classes
                   FROM users u LEFT JOIN staff_classes sc ON sc.staff_id=u.id
                   WHERE u.role IN ('staff','admin')
                   GROUP BY u.id ORDER BY u.name""")
    staff_list = cur.fetchall()

    # ── recent attendance (today, all classes) ──
    cur.execute("""SELECT a.*, u.name AS student_name, u.class_name,
                   st.name AS staff_name
                   FROM attendance a
                   JOIN users u ON u.id=a.student_id
                   JOIN users st ON st.id=a.staff_id
                   WHERE a.date=CURDATE()
                   ORDER BY a.marked_at DESC LIMIT 30""")
    recent_att = cur.fetchall()

    cur.close()
    return render_template('admin/dashboard.html',
        user=user, total_students=total_students, total_staff=total_staff,
        today_att=today_att, pending_certs=pending_certs,
        class_stats=class_stats, staff_list=staff_list, recent_att=recent_att,
        all_classes=ALL_CLASSES)

# ── Admin: view students of a class ──
@app.route('/admin/class/<class_name>')
@app.route('/admin/class/<class_name>/year/<int:year>')
@login_required
@admin_required
def admin_class_view(class_name, year=None):
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()

    # For 1st year classes (I-A etc.) year is always 1
    if class_name.startswith('I-'):
        year = 1

    params = [class_name]
    year_filter = ""
    if year:
        year_filter = " AND u.year=%s"
        params.append(year)

    cur.execute("""SELECT u.*,
                   (SELECT COUNT(*) FROM attendance WHERE student_id=u.id AND status='present') AS present_count,
                   (SELECT COUNT(*) FROM attendance WHERE student_id=u.id) AS total_att,
                   (SELECT COUNT(*) FROM certificates WHERE student_id=u.id AND verification_status='verified') AS certs,
                   (SELECT COUNT(*) FROM projects WHERE student_id=u.id) AS projects_count,
                   (SELECT COUNT(*) FROM submissions WHERE student_id=u.id) AS submissions_count
                   FROM users u WHERE u.class_name=%s AND u.role='student'""" + year_filter +
                   """ ORDER BY CAST(u.roll_number AS UNSIGNED), u.roll_number""", params)
    students = cur.fetchall()

    att_params = [class_name]
    att_year_filter = ""
    if year:
        att_year_filter = """ AND a.student_id IN (SELECT id FROM users WHERE class_name=%s AND year=%s)"""
        att_params = [class_name, class_name, year]

    cur.execute("""SELECT a.*, st.name AS staff_name
                   FROM attendance a JOIN users st ON st.id=a.staff_id
                   WHERE a.class_name=%s""" + att_year_filter +
                   """ ORDER BY a.date DESC, a.period LIMIT 100""", att_params)
    att_records = cur.fetchall()
    cur.close()
    return render_template('admin/class_view.html',
        user=user, class_name=class_name, year=year, students=students, att_records=att_records)

# ── Admin: Timetable management ──
@app.route('/admin/timetable', methods=['GET','POST'])
@login_required
@admin_required
def admin_timetable():
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()

    if request.method == 'POST':
        action = request.form.get('action','add')
        if action == 'add':
            staff_id   = request.form.get('staff_id', type=int)
            class_name = request.form.get('class_name','')
            year       = request.form.get('year', type=int, default=1)
            day        = request.form.get('day','')
            period     = request.form.get('period', type=int, default=1)
            subject    = request.form.get('subject','')
            start_time = request.form.get('start_time','')
            end_time   = request.form.get('end_time','')
            cur.execute("""INSERT INTO timetable (staff_id,class_name,year,day,period,subject,start_time,end_time)
                           VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
                        (staff_id,class_name,year,day,period,subject,start_time,end_time))
            mysql.connection.commit()
            flash('Timetable slot added ✓','success')
        elif action == 'bulk_save':
            class_name = request.form.get('class_name','')
            year       = request.form.get('year', type=int, default=1)
            # Delete existing slots for this class+year only
            if class_name:
                cur.execute("DELETE FROM timetable WHERE class_name=%s AND year=%s", (class_name, year))
            saved = 0
            days = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday']
            for day in days:
                for p in [1,2,3,4,5,6,7,8]:
                    subject    = request.form.get(f'slot_{day}_{p}_subject','').strip()
                    s_staff_id = request.form.get(f'slot_{day}_{p}_staff', type=int)
                    start_time = request.form.get(f'slot_{day}_{p}_start','')
                    end_time   = request.form.get(f'slot_{day}_{p}_end','')
                    if subject and s_staff_id and class_name:
                        cur.execute("""INSERT INTO timetable (staff_id,class_name,year,day,period,subject,start_time,end_time)
                                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
                                    (s_staff_id, class_name, year, day, p, subject, start_time, end_time))
                        saved += 1
            mysql.connection.commit()
            flash(f'Timetable saved with {saved} slots ✓','success')
        cur.close()
        redir_class = request.form.get('class_name','')
        redir_year  = request.form.get('year','')
        return redirect(url_for('admin_timetable', **({'class': redir_class, 'year': redir_year} if redir_class else {})))

    selected_class = request.args.get('class','')
    selected_year  = request.args.get('year', type=int)
    selected_staff = request.args.get('staff_id', type=int)
    selected_sem   = request.args.get('sem', type=int)
    draft_mode     = bool(request.args.get('draft',''))

    cur.execute("SELECT id,name,department FROM users WHERE role IN ('staff','admin') ORDER BY department, name")
    staff_list = cur.fetchall()

    # Fetch subjects from subject_master for the selected class/year so timetable can use them
    subjects_for_tt = []
    if selected_class and selected_year:
        # Determine dept from class name (e.g. AD-B -> AD, I-A -> first year no dept filter)
        cls_parts = selected_class.split('-')
        cls_dept  = cls_parts[0] if len(cls_parts) > 1 and not cls_parts[0].startswith('I') else ''
        sem_min   = (selected_year - 1) * 2 + 1
        sem_max   = sem_min + 1
        if cls_dept:
            cur.execute("""SELECT id, subject, subject_type, semester
                           FROM subject_master
                           WHERE department=%s AND semester IN (%s,%s)
                           ORDER BY semester, subject_type, subject""",
                        (cls_dept, sem_min, sem_max))
        else:
            # 1st year — fetch sem 1 & 2 regardless of dept
            cur.execute("""SELECT id, subject, subject_type, semester
                           FROM subject_master
                           WHERE semester IN (%s,%s)
                           ORDER BY semester, subject_type, subject""",
                        (sem_min, sem_max))
        subjects_for_tt = cur.fetchall()

    conditions = []
    params = []
    if selected_class:
        conditions.append("t.class_name=%s"); params.append(selected_class)
    if selected_year:
        # Match exact year OR rows saved before year column existed (year=1 default or NULL)
        conditions.append("(t.year=%s OR t.year IS NULL OR t.year=0)")
        params.append(selected_year)
    if selected_staff:
        conditions.append("t.staff_id=%s"); params.append(selected_staff)

    q = "SELECT t.*, u.name AS staff_name FROM timetable t JOIN users u ON u.id=t.staff_id"
    if conditions:
        q += " WHERE " + " AND ".join(conditions)
    q += " ORDER BY FIELD(t.day,'Monday','Tuesday','Wednesday','Thursday','Friday','Saturday'), t.period"

    cur.execute(q, params)
    timetable = cur.fetchall()
    cur.close()

    return render_template('admin/timetable.html',
        user=user, timetable=timetable, staff_list=staff_list,
        subjects_for_tt=subjects_for_tt,
        selected_class=selected_class, selected_year=selected_year,
        selected_staff=selected_staff, selected_sem=selected_sem,
        draft_mode=draft_mode, all_classes=ALL_CLASSES)

@app.route('/admin/timetable/delete/<int:slot_id>')
@login_required
@admin_required
def admin_delete_slot(slot_id):
    cur = get_cur()
    cur.execute("DELETE FROM timetable WHERE id=%s", (slot_id,))
    mysql.connection.commit(); cur.close()
    flash('Slot deleted','success')
    return redirect(url_for('admin_timetable'))

# ── Admin: attendance overview ──
@app.route('/admin/attendance')
@login_required
@admin_required
def admin_attendance():
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()

    sel_class = request.args.get('class','')
    sel_date  = request.args.get('date', str(date.today()))
    sel_year  = request.args.get('year', type=int)

    q = """SELECT a.*, u.name AS student_name, u.roll_number, u.class_name,
                  st.name AS staff_name, a.is_swap, a.swap_note
           FROM attendance a
           JOIN users u ON u.id=a.student_id
           JOIN users st ON st.id=a.staff_id
           WHERE a.date=%s"""
    params = [sel_date]
    if sel_class:
        q += " AND a.class_name=%s"
        params.append(sel_class)
    if sel_year:
        q += " AND u.year=%s"
        params.append(sel_year)
    q += " ORDER BY a.class_name, a.period, u.name"

    cur.execute(q, params)
    records = cur.fetchall()

    # summary per class
    sum_q = """SELECT class_name,
               SUM(status='present') AS present,
               SUM(status='absent') AS absent,
               SUM(status='late') AS late,
               COUNT(*) AS total
               FROM attendance WHERE date=%s"""
    sum_params = [sel_date]
    if sel_class:
        sum_q += " AND class_name=%s"; sum_params.append(sel_class)
    sum_q += " GROUP BY class_name ORDER BY class_name"
    cur.execute(sum_q, sum_params)
    summary = cur.fetchall()
    cur.close()

    return render_template('admin/attendance.html',
        user=user, records=records, summary=summary,
        sel_class=sel_class, sel_date=sel_date, sel_year=sel_year, all_classes=ALL_CLASSES)

# ── Admin: manage staff ──
@app.route('/admin/staff')
@login_required
@admin_required
def admin_staff():
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()
    cur.execute("""SELECT u.*,
                   GROUP_CONCAT(DISTINCT sc.class_name ORDER BY sc.class_name SEPARATOR ', ') AS classes,
                   GROUP_CONCAT(DISTINCT CASE WHEN sc.is_coordinator=1 THEN sc.coordinator_class END) AS cc_class
                   FROM users u LEFT JOIN staff_classes sc ON sc.staff_id=u.id
                   WHERE u.role IN ('staff','admin')
                   GROUP BY u.id ORDER BY u.name""")
    staff_list = cur.fetchall()
    cur.close()
    return render_template('admin/staff.html', user=user, staff_list=staff_list)


# ════════════════════════════════════════════
#  ADMIN SUBJECT MASTER (dept + year + sem + type)
# ════════════════════════════════════════════

DEPARTMENTS = ['CS','CE','MECH','AD','ECE','IT','S&H','T&P']
YEAR_SEM_MAP = {1:(1,2), 2:(3,4), 3:(5,6), 4:(7,8)}

@app.route('/admin/subjects', methods=['GET','POST'])
@login_required
@admin_required
def admin_subjects():
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()

    if request.method == 'POST':
        action = request.form.get('action','add')
        if action == 'add':
            dept    = request.form.get('department','').strip()
            year    = request.form.get('year', type=int)
            sem     = request.form.get('semester', type=int)
            subject = request.form.get('subject','').strip()
            stype   = request.form.get('subject_type','theory')
            if dept and sem and subject:
                cur.execute("""INSERT IGNORE INTO subject_master
                               (department, semester, subject, subject_type)
                               VALUES (%s,%s,%s,%s)""",
                            (dept, sem, subject, stype))
                mysql.connection.commit()
                flash(f'Subject "{subject}" added ✓','success')
            else:
                flash('Please fill department, semester and subject name.','warning')
        elif action == 'delete':
            sid = request.form.get('subject_id', type=int)
            cur.execute("DELETE FROM subject_master WHERE id=%s", (sid,))
            mysql.connection.commit()
            flash('Subject deleted','success')
        cur.close()
        return redirect(url_for('admin_subjects',
                                dept=request.form.get('department',''),
                                year=request.form.get('year','')))

    sel_dept = request.args.get('dept','')
    sel_year = request.args.get('year', type=int)

    # Build subject list filtered by dept + year (via semester range)
    q = "SELECT * FROM subject_master WHERE 1=1"
    params = []
    if sel_dept:
        q += " AND department=%s"; params.append(sel_dept)
    if sel_year:
        sem_min, sem_max = YEAR_SEM_MAP.get(sel_year, (1,8))
        q += " AND semester BETWEEN %s AND %s"
        params += [sem_min, sem_max]
    q += " ORDER BY department, semester, subject_type, subject"
    cur.execute(q, params)
    subjects = cur.fetchall()
    cur.close()

    return render_template('admin/subjects.html',
        user=user, subjects=subjects,
        DEPARTMENTS=DEPARTMENTS, YEAR_SEM_MAP=YEAR_SEM_MAP,
        sel_dept=sel_dept, sel_year=sel_year)


# ════════════════════════════════════════════
#  STAFF DASHBOARD — 4 class buttons
# ════════════════════════════════════════════

@app.route('/staff')
@login_required
@staff_required
def staff_dashboard():
    user = get_current_user()
    if not user: return redirect(url_for('login'))

    if session.get('role') == 'admin':
        return redirect(url_for('admin_dashboard'))

    staff_classes = get_staff_classes(user['id'])
    # active class from query param, default to first
    active_class = request.args.get('class') or (staff_classes[0] if staff_classes else None)
    return _staff_class_dashboard(user, staff_classes, active_class)

def _staff_class_dashboard(user, staff_classes, active_class):
    cur = get_cur()
    today = date.today()

    if not active_class:
        cur.close()
        return render_template('staff/dashboard.html',
            user=user, staff_classes=[], active_class=None,
            students=[], student_count=0, today_attendance={'marked':0},
            assignments=[], pending_certs=[], notifications=[],
            att_summary={}, quiz_stats=[], no_class=True)

    # Parse year prefix from active_class e.g. "3-AD-B" → year=3, cls="AD-B"
    cls_name, cls_year = parse_class_key(active_class)
    yr_filter, yr_params_suffix = year_filter_sql(cls_year, 'u')

    # Students
    cur.execute("""SELECT u.*,
                   (SELECT COUNT(*) FROM attendance WHERE student_id=u.id AND status='present') AS present_count,
                   (SELECT COUNT(*) FROM attendance WHERE student_id=u.id) AS total_att,
                   (SELECT COUNT(*) FROM certificates WHERE student_id=u.id AND verification_status='verified') AS certs,
                   (SELECT COUNT(*) FROM projects WHERE student_id=u.id) AS projects_count,
                   (SELECT COUNT(*) FROM submissions s2 JOIN assignments a2 ON a2.id=s2.assignment_id
                        WHERE s2.student_id=u.id AND a2.class_name=%s) AS submissions_count
                   FROM users u WHERE u.class_name=%s AND u.role='student'""" + yr_filter +
                   """ ORDER BY u.streak_stars DESC""",
                   [cls_name, cls_name] + yr_params_suffix)
    students = cur.fetchall()
    student_count = len(students)

    # Today attendance count
    cur.execute("""SELECT COUNT(DISTINCT student_id) AS marked,
                   SUM(status='present') AS present,
                   SUM(status='absent') AS absent,
                   SUM(status='late') AS late
                   FROM attendance
                   WHERE staff_id=%s AND date=%s AND class_name=%s""",
                (user['id'], today, cls_name))
    ta = cur.fetchone()
    today_attendance = ta if ta else {'marked':0,'present':0,'absent':0,'late':0}

    # Assignments for this class
    cur.execute("""SELECT a.*,
                   (SELECT COUNT(*) FROM submissions WHERE assignment_id=a.id) AS submitted_count,
                   (SELECT COUNT(*) FROM users WHERE class_name=a.class_name AND role='student') AS total_students
                   FROM assignments a WHERE a.staff_id=%s AND a.class_name=%s
                   ORDER BY a.created_at DESC LIMIT 8""", (user['id'], cls_name))
    assignments = cur.fetchall()

    # Pending certificates
    yr_cert_filter = " AND u.year=%s" if cls_year else ""
    cert_params = [cls_name] + yr_params_suffix
    cur.execute("""SELECT c.*, u.name AS student_name FROM certificates c
                   JOIN users u ON u.id=c.student_id
                   WHERE u.class_name=%s""" + yr_cert_filter + """ AND c.verification_status='pending'
                   ORDER BY c.created_at DESC LIMIT 5""", cert_params)
    pending_certs = cur.fetchall()

    # Quiz stats for this class
    cur.execute("""SELECT q.title, q.subject,
                   COUNT(qa.id) AS attempts,
                   AVG(qa.percentage) AS avg_score
                   FROM quizzes q LEFT JOIN quiz_attempts qa ON qa.quiz_id=q.id
                   WHERE q.staff_id=%s AND q.class_name=%s
                   GROUP BY q.id ORDER BY q.created_at DESC LIMIT 5""",
                (user['id'], cls_name))
    quiz_stats = cur.fetchall()

    # Notifications
    cur.execute("SELECT * FROM notifications WHERE user_id=%s ORDER BY created_at DESC LIMIT 5",
                (user['id'],))
    notifications = cur.fetchall()

    # Attendance summary per subject
    cur.execute("""SELECT subject,
                   SUM(status='present') AS present,
                   SUM(status='absent') AS absent,
                   COUNT(*) AS total
                   FROM attendance WHERE staff_id=%s AND class_name=%s
                   GROUP BY subject""", (user['id'], cls_name))
    att_rows = cur.fetchall()
    att_summary = {r['subject']: r for r in att_rows}

    cur.close()

    return render_template('staff/dashboard.html',
        user=user, today=today,
        staff_classes=staff_classes, active_class=active_class,
        cls_name=cls_name, cls_year=cls_year,
        students=students, student_count=student_count,
        today_attendance=today_attendance,
        assignments=assignments, pending_certs=pending_certs,
        quiz_stats=quiz_stats, notifications=notifications,
        att_summary=att_summary, no_class=False)

# ════════════════════════════════════════════
#  STAFF ATTENDANCE — Class → Period → OK → Students
# ════════════════════════════════════════════

@app.route('/staff/attendance', methods=['GET','POST'])
@login_required
@staff_required
def staff_attendance():
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()
    today = date.today()

    staff_classes = get_staff_classes(user['id'])

    # Step state from query params
    sel_class  = request.args.get('class','')
    sel_period = request.args.get('period', type=int)
    confirmed  = request.args.get('confirmed','0') == '1'

    if request.method == 'POST':
        att_date    = request.form.get('date', str(today))
        student_ids = request.form.getlist('student_ids[]')
        statuses    = request.form.getlist('statuses[]')
        class_name  = request.form.get('class_name','')
        subject     = request.form.get('subject','')
        period      = request.form.get('period', type=int, default=1)
        # If period is 99 (manual mode), use period_manual field instead
        if period == 99:
            period = request.form.get('period_manual', type=int, default=1)
        is_swap     = 1 if request.form.get('is_swap') else 0
        swap_note   = request.form.get('swap_note','')
        tt_id       = request.form.get('timetable_id') or None

        saved = 0
        present_count = 0
        for sid, status in zip(student_ids, statuses):
            try:
                # Check if this is a new record or an update
                cur.execute("SELECT id, status FROM attendance WHERE student_id=%s AND date=%s AND period=%s AND subject=%s",
                            (sid, att_date, period, subject))
                existing = cur.fetchone()

                cur.execute("""INSERT INTO attendance
                               (student_id,staff_id,class_name,subject,period,timetable_id,
                                date,status,is_swap,swap_note)
                               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                               ON DUPLICATE KEY UPDATE
                               status=%s, is_swap=%s, swap_note=%s, marked_at=NOW()""",
                            (sid, user['id'], class_name, subject, period, tt_id,
                             att_date, status, is_swap, swap_note,
                             status, is_swap, swap_note))
                saved += 1

                # ── Student stars for attendance ──
                # Award star only on NEW present record (not updates)
                if status == 'present' and not existing:
                    add_stars(int(sid), 2, f'Attendance: {subject} P{period}', 'attendance')
                    present_count += 1
                elif status == 'absent' and existing and existing['status'] == 'present':
                    # Was present, now marked absent — remove star
                    add_stars(int(sid), -2, f'Attendance corrected: {subject}', 'attendance')

            except Exception as e:
                print(f"Att insert error: {e}")

        # ── Staff gets 1 star per attendance session (flat) ──
        if saved > 0:
            mysql.connection.commit()
            add_stars(user['id'], 1, f'Marked attendance: {class_name} {subject} P{period}', 'attendance')
            flash(f'Attendance saved for {saved} students ✓  +1⭐ session star','success')
        else:
            mysql.connection.commit()
            flash(f'Attendance saved for {saved} students ✓','success')
        cur.close()
        return redirect(url_for('staff_attendance', **{'class': class_name}))

    # Fetch today's slots for selected class
    today_slots = []
    students    = []
    sel_slot    = None

    if sel_class:
        today_slots = get_today_slots(user['id'], sel_class)

    if sel_class and sel_period is not None and confirmed:
        # Find the timetable slot
        for sl in today_slots:
            if sl['period'] == sel_period:
                sel_slot = sl
                break

        # Parse year prefix e.g. "3-AD-B" → cls_name="AD-B", cls_year=3
        cls_name, cls_year = parse_class_key(sel_class)
        yf_sql, yf_params = year_filter_sql(cls_year)

        cur.execute("""SELECT id, name, roll_number FROM users
                       WHERE class_name=%s AND role='student'""" + yf_sql +
                       """ ORDER BY roll_number ASC""",
                    [cls_name] + yf_params)
        students = cur.fetchall()

    # Today's record for selected class
    today_att = []
    if sel_class:
        cls_name2, cls_year2 = parse_class_key(sel_class)
        cur.execute("""SELECT a.*, u.name AS student_name, u.roll_number
                       FROM attendance a JOIN users u ON u.id=a.student_id
                       WHERE a.staff_id=%s AND a.date=%s AND a.class_name=%s
                       ORDER BY a.period, u.name""",
                    (user['id'], today, cls_name2))
        today_att = cur.fetchall()

    cur.close()
    return render_template('staff/attendance.html',
        user=user, today=today,
        staff_classes=staff_classes,
        sel_class=sel_class, sel_period=sel_period, confirmed=confirmed,
        today_slots=today_slots, sel_slot=sel_slot,
        students=students, today_att=today_att)

# ── API: get periods for a class (for AJAX) ──
@app.route('/api/periods')
@login_required
def api_periods():
    class_name = request.args.get('class','')
    staff_id   = session.get('user_id')
    day_name   = date.today().strftime('%A')
    cur = get_cur()
    cur.execute("""SELECT * FROM timetable WHERE staff_id=%s AND class_name=%s AND day=%s
                   ORDER BY period""", (staff_id, class_name, day_name))
    slots = cur.fetchall()
    cur.close()
    result = []
    for s in slots:
        result.append({
            'id': s['id'],
            'period': s['period'],
            'subject': s['subject'],
            'start_time': str(s['start_time']),
            'end_time': str(s['end_time'])
        })
    return jsonify(result)

# ════════════════════════════════════════════
#  STAFF PROFILE — checkbox class selection
# ════════════════════════════════════════════

@app.route('/profile', methods=['GET','POST'])
@login_required
def profile():
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()

    if request.method == 'POST':
        phone = request.form.get('phone','')
        pic   = request.files.get('profile_pic')
        pic_path = user.get('profile_pic')

        if pic and allowed(pic.filename, ALLOWED_IMG):
            pic_path = save_file(pic, 'profiles')

        cur.execute("UPDATE users SET phone=%s, profile_pic=%s WHERE id=%s",
                    (phone, pic_path, user['id']))

        # Staff: update class assignments (max 5) + CC
        if user['role'] in ('staff','admin'):
            selected = request.form.getlist('staff_classes')[:5]
            is_cc    = 1 if request.form.get('is_coordinator') else 0
            cc_class = request.form.get('coordinator_class','')
            cc_year  = request.form.get('coordinator_year','')
            cc_key   = f"{cc_year}-{cc_class}" if (is_cc and cc_year and cc_class) else ''

            cur.execute("DELETE FROM staff_classes WHERE staff_id=%s", (user['id'],))
            for cls in selected:
                if cls in ALL_CLASSES:
                    cur.execute("INSERT IGNORE INTO staff_classes (staff_id,class_name) VALUES (%s,%s)",
                                (user['id'], cls))
            # Save CC info
            if is_cc and cc_key:
                cur.execute("""INSERT INTO staff_classes (staff_id,class_name,is_coordinator,coordinator_class)
                               VALUES (%s,%s,1,%s)
                               ON DUPLICATE KEY UPDATE is_coordinator=1, coordinator_class=%s""",
                            (user['id'], cc_key, cc_key, cc_key))

        mysql.connection.commit()
        flash('Profile updated ✓','success')
        cur.close()
        return redirect(url_for('profile'))

    cur.execute("SELECT * FROM streak_log WHERE user_id=%s ORDER BY created_at DESC LIMIT 30",
                (user['id'],))
    streak_history = cur.fetchall()

    my_classes = get_staff_classes(user['id']) if user['role'] in ('staff','admin') else []

    # Get CC info for staff
    cc_info = None
    if user['role'] in ('staff','admin'):
        try:
            cur.execute("""SELECT coordinator_class, is_coordinator FROM staff_classes
                           WHERE staff_id=%s AND is_coordinator=1 LIMIT 1""", (user['id'],))
            cc_row = cur.fetchone()
            if cc_row and cc_row['coordinator_class']:
                raw = cc_row['coordinator_class']
                # Parse "3-AD-B" → year=3, class=AD-B  or "I-A" → year=1, class=I-A
                import re as _re
                m = _re.match(r'^(\d+)-(.+)$', raw)
                if m:
                    cc_info = {'cc_year': int(m.group(1)), 'cc_class': m.group(2), 'raw': raw}
                else:
                    cc_info = {'cc_year': 1, 'cc_class': raw, 'raw': raw}
        except:
            pass

    cur.close()

    return render_template('profile.html',
        user=user, streak_history=streak_history,
        my_classes=my_classes, cc_info=cc_info, all_classes=ALL_CLASSES)

# ════════════════════════════════════════════
#  TIMETABLE (staff view only — admin manages)
# ════════════════════════════════════════════

@app.route('/staff/timetable')
@login_required
@staff_required
def staff_timetable():
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    if session.get('role') == 'admin':
        return redirect(url_for('admin_timetable'))
    cur = get_cur()
    cur.execute("""SELECT * FROM timetable WHERE staff_id=%s
                   ORDER BY FIELD(day,'Monday','Tuesday','Wednesday','Thursday','Friday','Saturday'),period""",
                (user['id'],))
    timetable = cur.fetchall()
    cur.close()
    return render_template('staff/timetable.html', user=user, timetable=timetable)

@app.route('/admin/timetable/slot/delete/<int:slot_id>')
@login_required
@admin_required
def delete_timetable_slot(slot_id):
    cur = get_cur()
    cur.execute("DELETE FROM timetable WHERE id=%s", (slot_id,))
    mysql.connection.commit(); cur.close()
    flash('Slot deleted','success')
    return redirect(url_for('admin_timetable'))

# ════════════════════════════════════════════
#  ASSIGNMENTS
# ════════════════════════════════════════════

@app.route('/staff/assignments', methods=['GET','POST'])
@login_required
@staff_required
def staff_assignments():
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()
    today = date.today()
    staff_classes = get_staff_classes(user['id'])

    if request.method == 'POST':
        title    = request.form.get('title','')
        desc     = request.form.get('description','')
        subject  = request.form.get('subject','')
        due_date = request.form.get('due_date','')
        max_m    = request.form.get('max_marks', 10, type=int)
        target_raw = request.form.get('target_classes') or request.form.getlist('target_classes')
        if isinstance(target_raw, str):
            target = [target_raw] if target_raw else staff_classes
        else:
            target = target_raw or staff_classes

        for cls in target:
            cls_name, cls_year = parse_class_key(cls)
            yf_sql, yf_params = year_filter_sql(cls_year)
            cur.execute("""INSERT INTO assignments (staff_id,class_name,title,description,subject,due_date,max_marks)
                           VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                        (user['id'], cls_name, title, desc, subject, due_date, max_m))
            cur.execute("SELECT id FROM users WHERE class_name=%s AND role='student'" + yf_sql,
                        [cls_name] + yf_params)
            for s in cur.fetchall():
                _notif(cur, s['id'], f'New Assignment: {title}',
                       f'{subject} — Due {due_date}', 'assignment')
        mysql.connection.commit()
        add_stars(user['id'], 3, f'Posted assignment: {title}', 'assignment')
        flash('Assignment posted ✓  +3⭐','success')
        cur.close()
        return redirect(url_for('staff_assignments'))

    active_class = request.args.get('class', '')
    cls_name, cls_year = parse_class_key(active_class) if active_class else (None, None)

    if active_class and cls_name:
        cur.execute("""SELECT a.*,
                       (SELECT COUNT(*) FROM submissions WHERE assignment_id=a.id) AS submitted_count,
                       (SELECT COUNT(*) FROM users WHERE class_name=a.class_name AND role='student') AS total_students
                       FROM assignments a WHERE a.staff_id=%s AND a.class_name=%s
                       ORDER BY a.created_at DESC""", (user['id'], cls_name))
    else:
        # Show all assignments across all staff classes
        all_cls_names = [parse_class_key(c)[0] for c in staff_classes]
        if all_cls_names:
            placeholders = ','.join(['%s'] * len(all_cls_names))
            cur.execute(f"""SELECT a.*,
                           (SELECT COUNT(*) FROM submissions WHERE assignment_id=a.id) AS submitted_count,
                           (SELECT COUNT(*) FROM users WHERE class_name=a.class_name AND role='student') AS total_students
                           FROM assignments a WHERE a.staff_id=%s AND a.class_name IN ({placeholders})
                           ORDER BY a.created_at DESC""", [user['id']] + all_cls_names)
        else:
            cur.execute("""SELECT a.*,
                           (SELECT COUNT(*) FROM submissions WHERE assignment_id=a.id) AS submitted_count,
                           0 AS total_students
                           FROM assignments a WHERE a.staff_id=%s
                           ORDER BY a.created_at DESC""", (user['id'],))
    assignments = cur.fetchall()
    cur.close()
    return render_template('staff/assignments.html',
        user=user, today=today, assignments=assignments,
        staff_classes=staff_classes, active_class=active_class)

@app.route('/student/assignments', methods=['GET','POST'])
@login_required
@student_required
def student_assignments():
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()
    today = date.today()

    if request.method == 'POST':
        assign_id = request.form.get('assignment_id', type=int)
        file      = request.files.get('file')
        cur.execute("SELECT * FROM assignments WHERE id=%s", (assign_id,))
        asn = cur.fetchone()
        if not asn: flash('Not found','danger'); cur.close(); return redirect(url_for('student_assignments'))
        cur.execute("SELECT id FROM submissions WHERE assignment_id=%s AND student_id=%s",
                    (assign_id, user['id']))
        if cur.fetchone(): flash('Already submitted','danger'); cur.close(); return redirect(url_for('student_assignments'))
        is_late = (date.fromisoformat(str(asn['due_date'])) < today)
        path    = save_file(file,'assignments') if (file and allowed(file.filename,ALLOWED_PDF)) else None
        cur.execute("INSERT INTO submissions (assignment_id,student_id,file_path,status) VALUES (%s,%s,%s,%s)",
                    (assign_id, user['id'], path, 'late' if is_late else 'submitted'))
        mysql.connection.commit()
        stars = app.config['STARS']['assignment_late'] if is_late else app.config['STARS']['assignment_ontime']
        add_stars(user['id'], stars, f'Assignment: {asn["title"]}', 'assignment')
        flash('Submitted! ✓','success')
        cur.close()
        return redirect(url_for('student_assignments'))

    cur.execute("""SELECT a.*, s.id AS submitted, s.status AS sub_status,
                   s.marks_obtained AS marks, s.feedback
                   FROM assignments a
                   LEFT JOIN submissions s ON s.assignment_id=a.id AND s.student_id=%s
                   WHERE a.class_name=%s ORDER BY a.due_date""",
                (user['id'], user['class_name']))
    assignments = cur.fetchall()
    cur.close()
    return render_template('student/assignments.html', user=user, today=today, assignments=assignments)

@app.route('/staff/submissions/<int:assign_id>')
@login_required
@staff_required
def view_submissions(assign_id):
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()
    today = date.today()
    cur.execute("SELECT * FROM assignments WHERE id=%s", (assign_id,))
    assignment = cur.fetchone()
    if not assignment: flash('Not found','danger'); cur.close(); return redirect(url_for('staff_assignments'))
    cur.execute("""SELECT s.*, u.name AS student_name, u.roll_number
                   FROM submissions s JOIN users u ON u.id=s.student_id
                   WHERE s.assignment_id=%s ORDER BY s.submitted_at""", (assign_id,))
    submissions = cur.fetchall()
    submitted_ids = [s['student_id'] for s in submissions]
    cur.execute("SELECT * FROM users WHERE class_name=%s AND role='student'", (assignment['class_name'],))
    not_submitted = [s for s in cur.fetchall() if s['id'] not in submitted_ids]
    cur.close()
    return render_template('staff/submissions.html',
        user=user, assignment=assignment, submissions=submissions,
        not_submitted=not_submitted, today=today)

@app.route('/staff/grade/<int:sub_id>', methods=['POST'])
@login_required
@staff_required
def grade_submission(sub_id):
    cur = get_cur()
    marks    = request.form.get('marks', type=float)
    feedback = request.form.get('feedback','')
    cur.execute("""SELECT s.*, a.max_marks, a.title FROM submissions s
                   JOIN assignments a ON a.id=s.assignment_id WHERE s.id=%s""", (sub_id,))
    sub = cur.fetchone()
    if sub:
        cur.execute("UPDATE submissions SET marks_obtained=%s,feedback=%s,status='graded' WHERE id=%s",
                    (marks, feedback, sub_id))
        mysql.connection.commit()
        if marks and sub['max_marks']:
            pct = (marks/sub['max_marks'])*100
            stars = (app.config['STARS']['result_above75'] if pct>=75 else
                     app.config['STARS']['result_above60'] if pct>=60 else
                     app.config['STARS']['result_above40'] if pct>=40 else 0)
            if stars: add_stars(sub['student_id'], stars, f'Assignment graded: {sub["title"]}')
    cur.close()
    flash('Graded ✓','success')
    return redirect(request.referrer or url_for('staff_assignments'))

# ════════════════════════════════════════════
#  QUIZ
# ════════════════════════════════════════════

# ════════════════════════════════════════════
#  QUIZ — v2 (multi-type questions)
# ════════════════════════════════════════════

@app.route('/staff/quiz', methods=['GET','POST'])
@login_required
@staff_required
def staff_quiz():
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()
    today = date.today()
    staff_classes = get_staff_classes(user['id'])

    if request.method == 'POST' and request.form.get('action') == 'create_quiz':
        title   = request.form.get('title','')
        subject = request.form.get('subject','')
        dur     = request.form.get('duration_minutes', 30, type=int)
        due_dt  = request.form.get('due_date') or None
        qs_json = request.form.get('questions_json','[]')
        questions = json.loads(qs_json)
        total_marks = sum(int(q.get('marks',1)) for q in questions)
        target = request.form.getlist('target_classes') or staff_classes

        for cls in target:
            cls_name, cls_year = parse_class_key(cls)
            yf_sql, yf_params = year_filter_sql(cls_year)
            cur.execute("""INSERT INTO quizzes
                           (staff_id,class_name,title,subject,duration_minutes,total_marks,due_date)
                           VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                        (user['id'], cls_name, title, subject, dur, total_marks, due_dt))
            quiz_id = cur.lastrowid
            for q in questions:
                cur.execute("""INSERT INTO quiz_questions_v2
                               (quiz_id,question,q_type,options_json,correct_json,marks)
                               VALUES (%s,%s,%s,%s,%s,%s)""",
                            (quiz_id, q['text'], q['type'],
                             json.dumps(q.get('options',[])),
                             json.dumps(q.get('correct',[])),
                             int(q.get('marks',1))))
            cur.execute("SELECT id FROM users WHERE class_name=%s AND role='student'" + yf_sql,
                        [cls_name] + yf_params)
            for s in cur.fetchall():
                _notif(cur, s['id'], f'New Quiz: {title}', f'{subject}', 'quiz')
        mysql.connection.commit()
        flash('Quiz created ✓','success')
        cur.close()
        return redirect(url_for('staff_quiz'))

    active_class = request.args.get('class') or (staff_classes[0] if staff_classes else '')
    cls_name, cls_year = parse_class_key(active_class)
    cur.execute("""SELECT q.*,
                   (SELECT COUNT(*) FROM quiz_attempts WHERE quiz_id=q.id) AS attempts,
                   (SELECT AVG(percentage) FROM quiz_attempts WHERE quiz_id=q.id) AS avg_score
                   FROM quizzes q WHERE q.staff_id=%s AND q.class_name=%s
                   ORDER BY q.created_at DESC""", (user['id'], cls_name))
    quizzes = cur.fetchall()
    cur.close()
    return render_template('staff/quiz.html',
        user=user, today=today, quizzes=quizzes,
        staff_classes=staff_classes, active_class=active_class)

@app.route('/staff/quiz/<int:quiz_id>/results')
@login_required
@staff_required
def quiz_results(quiz_id):
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()
    cur.execute("SELECT * FROM quizzes WHERE id=%s", (quiz_id,))
    quiz = cur.fetchone()
    cur.execute("""SELECT a.*, u.name AS student_name, u.roll_number,
                   (SELECT COUNT(*) FROM quiz_answers qa
                    JOIN quiz_questions_v2 qq ON qq.id=qa.question_id
                    WHERE qa.attempt_id=a.id AND qq.q_type IN ('short_answer','long_answer') AND qa.is_graded=0) AS pending_grading
                   FROM quiz_attempts a JOIN users u ON u.id=a.student_id
                   WHERE a.quiz_id=%s ORDER BY a.score DESC""", (quiz_id,))
    attempts = cur.fetchall()
    # Add grade to each attempt
    for a in attempts:
        a['grade'] = calc_grade(a['percentage'] or 0)
    cur.close()
    return render_template('staff/quiz_results.html', user=user, quiz=quiz, attempts=attempts)

@app.route('/staff/quiz/<int:quiz_id>/grade', methods=['GET','POST'])
@login_required
@staff_required
def staff_grade_quiz(quiz_id):
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()
    cur.execute("SELECT * FROM quizzes WHERE id=%s", (quiz_id,))
    quiz = cur.fetchone()

    if request.method == 'POST':
        answer_id  = request.form.get('answer_id', type=int)
        marks_given = request.form.get('marks_given', type=float, default=0)
        cur.execute("UPDATE quiz_answers SET marks_given=%s, is_graded=1 WHERE id=%s",
                    (marks_given, answer_id))
        # Recalculate attempt score
        cur.execute("SELECT attempt_id FROM quiz_answers WHERE id=%s", (answer_id,))
        row = cur.fetchone()
        if row:
            cur.execute("SELECT SUM(marks_given) AS total FROM quiz_answers WHERE attempt_id=%s",
                        (row['attempt_id'],))
            score_row = cur.fetchone()
            new_score = score_row['total'] or 0
            cur.execute("SELECT total_marks FROM quiz_attempts WHERE id=%s", (row['attempt_id'],))
            tm = cur.fetchone()
            total = tm['total_marks'] or 1
            pct = (new_score/total)*100
            cur.execute("UPDATE quiz_attempts SET score=%s,percentage=%s WHERE id=%s",
                        (new_score, pct, row['attempt_id']))
        mysql.connection.commit()
        flash('Marks saved ✓','success')
        cur.close()
        return redirect(url_for('staff_grade_quiz', quiz_id=quiz_id))

    # Fetch all text answers needing grading
    cur.execute("""SELECT qa.id AS answer_id, qa.answer_given, qa.marks_given, qa.is_graded,
                   qq.question, qq.q_type, qq.marks,
                   u.name AS student_name, u.roll_number
                   FROM quiz_answers qa
                   JOIN quiz_questions_v2 qq ON qq.id=qa.question_id
                   JOIN quiz_attempts at2 ON at2.id=qa.attempt_id
                   JOIN users u ON u.id=at2.student_id
                   WHERE at2.quiz_id=%s AND qq.q_type IN ('short_answer','long_answer')
                   ORDER BY qa.is_graded ASC, u.name ASC""", (quiz_id,))
    pending = cur.fetchall()
    cur.close()
    return render_template('staff/quiz_grade.html', user=user, quiz=quiz, pending=pending)

@app.route('/quiz/<int:quiz_id>/attempt', methods=['GET','POST'])
@login_required
@student_required
def attempt_quiz(quiz_id):
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()
    cur.execute("SELECT id FROM quiz_attempts WHERE quiz_id=%s AND student_id=%s", (quiz_id, user['id']))
    if cur.fetchone():
        flash('Already attempted','warning'); cur.close()
        return redirect(url_for('dashboard'))
    cur.execute("SELECT * FROM quizzes WHERE id=%s", (quiz_id,))
    quiz = cur.fetchone()
    cur.execute("SELECT * FROM quiz_questions_v2 WHERE quiz_id=%s ORDER BY id", (quiz_id,))
    questions = cur.fetchall()

    if request.method == 'POST':
        # Create attempt first
        cur.execute("""INSERT INTO quiz_attempts (quiz_id,student_id,score,total_marks,percentage)
                       VALUES (%s,%s,0,%s,0)""", (quiz_id, user['id'], quiz['total_marks'] or 0))
        attempt_id = cur.lastrowid
        auto_score = 0

        for q in questions:
            qtype = q['q_type']
            opts  = json.loads(q['options_json'] or '[]')
            corr  = json.loads(q['correct_json'] or '[]')
            marks = q['marks']
            given_marks = 0
            is_graded = 1

            if qtype == 'mcq':
                ans = request.form.get(f'q_{q["id"]}','')
                answer_str = ans
                try:
                    given_marks = marks if int(ans) in corr else 0
                except: given_marks = 0

            elif qtype == 'multi_select':
                ans_list = request.form.getlist(f'q_{q["id"]}[]')
                answer_str = json.dumps([int(a) for a in ans_list])
                try:
                    sel = set(int(a) for a in ans_list)
                    given_marks = marks if sel == set(corr) else 0
                except: given_marks = 0

            else:  # short/long answer
                answer_str = request.form.get(f'q_{q["id"]}','').strip()
                is_graded  = 0  # needs manual grading

            auto_score += given_marks
            cur.execute("""INSERT INTO quiz_answers
                           (attempt_id,question_id,answer_given,marks_given,is_graded)
                           VALUES (%s,%s,%s,%s,%s)""",
                        (attempt_id, q['id'], answer_str, given_marks, is_graded))

        total = quiz['total_marks'] or 1
        pct   = (auto_score/total)*100
        cur.execute("UPDATE quiz_attempts SET score=%s,percentage=%s WHERE id=%s",
                    (auto_score, pct, attempt_id))
        mysql.connection.commit()
        add_stars(user['id'], app.config['STARS']['quiz_attempt'], f'Quiz: {quiz["title"]}', 'quiz')
        flash(f'Submitted! Auto-scored: {auto_score}/{total}. Text answers will be graded by staff.','success')
        cur.close()
        return redirect(url_for('dashboard'))

    cur.close()
    return render_template('student/quiz.html', user=user, quiz=quiz, questions=questions)

@app.route('/staff/quiz/<int:quiz_id>/export')
@login_required
@staff_required
def export_quiz_excel(quiz_id):
    if not EXCEL_OK:
        flash('pip install openpyxl','danger')
        return redirect(url_for('quiz_results', quiz_id=quiz_id))
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    cur = get_cur()
    cur.execute("SELECT * FROM quizzes WHERE id=%s", (quiz_id,))
    quiz = cur.fetchone()

    cur.execute("SELECT * FROM quiz_questions_v2 WHERE quiz_id=%s ORDER BY id", (quiz_id,))
    questions = cur.fetchall()

    cur.execute("""SELECT a.*, u.name AS student_name, u.roll_number
                   FROM quiz_attempts a JOIN users u ON u.id=a.student_id
                   WHERE a.quiz_id=%s ORDER BY CAST(u.roll_number AS UNSIGNED)""", (quiz_id,))
    attempts = cur.fetchall()

    cur.execute("""SELECT qa.attempt_id, qa.question_id, qa.answer_given, qa.marks_given
                   FROM quiz_answers qa
                   JOIN quiz_attempts at2 ON at2.id=qa.attempt_id
                   WHERE at2.quiz_id=%s""", (quiz_id,))
    all_answers = cur.fetchall()
    cur.close()

    # Build answer lookup {attempt_id: {question_id: row}}
    ans_map = {}
    for row in all_answers:
        ans_map.setdefault(row['attempt_id'], {})[row['question_id']] = row

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Responses'

    # Styles
    thin        = Side(style='thin', color='CCCCCC')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wrap_top    = Alignment(wrap_text=True, vertical='top')

    hdr_font    = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill    = PatternFill('solid', fgColor='1F4E79')
    qhdr_fill   = PatternFill('solid', fgColor='2E75B6')
    qtype_fill  = PatternFill('solid', fgColor='BDD7EE')
    qtype_font  = Font(bold=True, color='1F4E79', size=9)

    # Row 1: headers
    fixed_headers = ['#', 'Student Name', 'Roll No', 'Score', 'Total', '%', 'Grade']
    q_headers = []
    for idx, q in enumerate(questions):
        label = f"Q{idx+1}: {q['question'][:60]}{'...' if len(q['question'])>60 else ''}"
        q_headers.append(label)
    all_headers = fixed_headers + q_headers
    ws.append(all_headers)

    for col, _ in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font      = hdr_font
        cell.fill      = hdr_fill if col <= len(fixed_headers) else qhdr_fill
        cell.alignment = center
        cell.border    = border

    # Row 2: question type + marks
    type_row = ['', '', '', '', '', '', '']
    for q in questions:
        qtype = q['q_type'].replace('_', ' ').title()
        type_row.append(f"{qtype}  |  {q['marks']} mark{'s' if q['marks']!=1 else ''}")
    ws.append(type_row)
    for col in range(len(fixed_headers)+1, len(all_headers)+1):
        cell = ws.cell(row=2, column=col)
        cell.font      = qtype_font
        cell.fill      = qtype_fill
        cell.alignment = center
        cell.border    = border

    # Data rows
    for i, attempt in enumerate(attempts, 1):
        aids = ans_map.get(attempt['id'], {})
        pct  = attempt['percentage'] or 0
        row  = [
            i,
            attempt['student_name'],
            attempt['roll_number'],
            attempt['score'],
            attempt['total_marks'],
            round(pct, 1),
            calc_grade(pct)
        ]

        for q in questions:
            ans_row = aids.get(q['id'])
            if not ans_row:
                row.append('—')
                continue
            raw = ans_row['answer_given'] or ''
            if q['q_type'] == 'mcq':
                try:
                    opts = json.loads(q['options_json'] or '[]')
                    raw  = opts[int(raw)] if int(raw) < len(opts) else raw
                except: pass
            elif q['q_type'] == 'multi_select':
                try:
                    opts    = json.loads(q['options_json'] or '[]')
                    indices = json.loads(raw)
                    raw     = ', '.join(opts[j] for j in indices if j < len(opts))
                except: pass
            row.append(f"{raw}  [{ans_row['marks_given']}/{q['marks']}]")

        ws.append(row)
        data_row = ws.max_row
        score_color = ('C6EFCE' if pct >= 75 else 'FFEB9C' if pct >= 50 else 'FFC7CE')
        ws.cell(row=data_row, column=6).fill = PatternFill('solid', fgColor=score_color)
        ws.cell(row=data_row, column=7).font = Font(bold=True)
        for col in range(1, len(all_headers)+1):
            cell = ws.cell(row=data_row, column=col)
            cell.alignment = wrap_top
            cell.border    = border

    # Column widths
    for col, w in enumerate([5,22,16,8,8,8,8], 1):
        ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width = w
    for col in range(len(fixed_headers)+1, len(all_headers)+1):
        ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width = 32

    ws.row_dimensions[1].height = 38
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = 'H3'

    path = os.path.join(TEMP_DIR, f'quiz_{quiz_id}.xlsx')
    wb.save(path)
    return send_file(path, as_attachment=True,
                     download_name=f'{quiz["title"]}_responses.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ════════════════════════════════════════════
#  RESULTS — CAT marks (staff) + Semester grades (student)
# ════════════════════════════════════════════

@app.route('/staff/results', methods=['GET','POST'])
@login_required
@staff_required
def staff_results():
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()
    staff_classes = get_staff_classes(user['id'])

    if request.method == 'POST':
        subject    = request.form.get('subject','').strip()
        semester   = request.form.get('semester', type=int)
        raw_class  = request.form.get('class_name','')
        cls_name, cls_year = parse_class_key(raw_class)
        class_name = cls_name  # store plain class name e.g. "AD-B" not "3-AD-B"

        if subject and semester and class_name:
            saved = 0
            for key, val in request.form.items():
                if key.startswith('cat1_') or key.startswith('cat2_'):
                    continue  # handled below
            # Collect all student ids present in form
            student_ids = set()
            for key in request.form.keys():
                if key.startswith('cat1_'):
                    try: student_ids.add(int(key[5:]))
                    except: pass
                elif key.startswith('cat2_'):
                    try: student_ids.add(int(key[5:]))
                    except: pass
            for sid in student_ids:
                cat1_str = request.form.get(f'cat1_{sid}','').strip()
                cat2_str = request.form.get(f'cat2_{sid}','').strip()
                cat1 = float(cat1_str) if cat1_str != '' else None
                cat2 = float(cat2_str) if cat2_str != '' else None
                if cat1 is None and cat2 is None:
                    continue  # skip blank rows
                cat1 = cat1 if cat1 is not None else 0
                cat2 = cat2 if cat2 is not None else 0
                # Delete existing then insert fresh — guaranteed upsert
                cur.execute("""DELETE FROM cat_marks
                               WHERE student_id=%s AND subject=%s AND semester=%s""",
                            (sid, subject, semester))
                cur.execute("""INSERT INTO cat_marks
                               (student_id,staff_id,class_name,subject,semester,cat1_marks,cat2_marks)
                               VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                            (sid, user['id'], class_name, subject, semester, cat1, cat2))
                saved += 1
            mysql.connection.commit()
            flash(f'CAT marks saved for {saved} students ✓','success')
        else:
            flash('Please enter subject and semester.','warning')
        cur.close()
        # Redirect back with subject+semester so marks show pre-filled
        redir_params = {'class': class_name}
        if subject:  redir_params['subject']  = subject
        if semester: redir_params['semester'] = semester
        return redirect(url_for('staff_results', **redir_params))

    active_class  = request.args.get('class') or (staff_classes[0] if staff_classes else '')
    filter_subject = request.args.get('subject','').strip()
    filter_sem     = request.args.get('semester', type=int)
    students = []
    cat_records = []
    if active_class:
        cls_name, cls_year = parse_class_key(active_class)
        yf_sql, yf_params = year_filter_sql(cls_year)
        cur.execute("""SELECT id,name,roll_number FROM users
                       WHERE class_name=%s AND role='student'""" + yf_sql +
                       """ ORDER BY roll_number""",
                    [cls_name] + yf_params)
        students = cur.fetchall()
        q2 = """SELECT cm.*, u.name AS student_name, u.roll_number, u.id AS student_id
                   FROM cat_marks cm JOIN users u ON u.id=cm.student_id
                   WHERE cm.class_name=%s"""
        p2 = [cls_name]
        if filter_subject:
            q2 += " AND cm.subject=%s"; p2.append(filter_subject)
        if filter_sem:
            q2 += " AND cm.semester=%s"; p2.append(filter_sem)
        q2 += " ORDER BY u.roll_number, cm.semester, cm.subject"
        cur.execute(q2, p2)
        raw_records = cur.fetchall()
        cat_lookup = {r['student_id']: r for r in raw_records}
        cat_records = raw_records
    else:
        cat_lookup = {}
    cur.close()
    return render_template('staff/results.html',
        user=user, students=students, cat_records=cat_records,
        cat_lookup=cat_lookup,
        staff_classes=staff_classes, active_class=active_class,
        filter_subject=filter_subject, filter_sem=filter_sem)

@app.route('/student/results', methods=['GET','POST'])
@login_required
@student_required
def student_results():
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()

    if request.method == 'POST':
        action = request.form.get('action','')
        if action == 'add_result':
            semester     = request.form.get('semester', type=int)
            subject      = request.form.get('subject','').strip()
            subject_type = request.form.get('subject_type','theory')
            grade        = request.form.get('grade','')
            if semester and subject and grade:
                cur.execute("""INSERT INTO semester_results
                               (student_id,semester,subject,subject_type,grade)
                               VALUES (%s,%s,%s,%s,%s)
                               ON DUPLICATE KEY UPDATE grade=%s""",
                            (user['id'], semester, subject, subject_type, grade, grade))
                mysql.connection.commit()

                # Notify CC of this student's class
                cc = get_class_coordinator(user.get('class_name',''))
                if cc:
                    _notif(cur, cc['id'],
                           f'Result Updated: {user["name"]}',
                           f'Sem {semester} — {subject}: {grade}',
                           'achievement')

                # Notify subject staff if assigned
                cur.execute("""SELECT sa.staff_id FROM subject_assignments sa
                               JOIN subject_master sm ON sm.id=sa.subject_id
                               WHERE sm.subject=%s AND sm.semester=%s AND sa.class_name=%s
                               LIMIT 1""",
                            (subject, semester, user.get('class_name','')))
                sa_row = cur.fetchone()
                if sa_row and (not cc or sa_row['staff_id'] != cc['id']):
                    _notif(cur, sa_row['staff_id'],
                           f'Result: {user["name"]}',
                           f'Sem {semester} — {subject}: {grade}',
                           'achievement')

                mysql.connection.commit()
                flash('Result added ✓','success')

        elif action == 'delete_result':
            result_id = request.form.get('result_id', type=int)
            cur.execute("DELETE FROM semester_results WHERE id=%s AND student_id=%s",
                        (result_id, user['id']))
            mysql.connection.commit()
            flash('Deleted','success')
        cur.close()
        return redirect(url_for('student_results'))

    cur.execute("""SELECT * FROM semester_results WHERE student_id=%s
                   ORDER BY semester, subject_type, subject""", (user['id'],))
    sem_results = cur.fetchall()

    cur.execute("""SELECT * FROM cat_marks WHERE student_id=%s
                   ORDER BY semester, subject""", (user['id'],))
    cat_marks = cur.fetchall()

    # Subjects from master for this student's department (for JS dropdown)
    dept = user.get('department','')
    cur.execute("""SELECT DISTINCT semester FROM subject_master
                   WHERE department=%s ORDER BY semester""", (dept,))
    available_sems = [r['semester'] for r in cur.fetchall()]

    cur.close()
    return render_template('student/results.html',
        user=user, sem_results=sem_results, cat_marks=cat_marks,
        available_sems=available_sems)

# ════════════════════════════════════════════
#  CERTIFICATES
# ════════════════════════════════════════════

CERT_STARS = {'workshop_attended':10,'workshop_conducted':12,'sports':12,'paper_presentation':10,'project_expo':10,
              'inter_college':15,'intra_college':10,'seminar':8,'club_event':8,'other':5}

@app.route('/student/certificates', methods=['GET','POST'])
@login_required
@student_required
def student_certificates():
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()
    if request.method == 'POST':
        title=request.form.get('title','').strip(); category=request.form.get('category','other')
        event_title=request.form.get('event_title','').strip()
        issue_date=request.form.get('issue_date') or None; issuer=request.form.get('issuer','').strip()
        file=request.files.get('file'); path=None; ai_conf=0.0
        if file and allowed(file.filename,ALLOWED_CERT):
            path=save_file(file,'certs')
            ai_conf=0.80 if file.filename.lower().endswith('.pdf') else 0.65
        vstatus='verified' if ai_conf>=0.75 else ('manual_review' if ai_conf>=0.5 else 'pending')
        stars=CERT_STARS.get(category,5) if vstatus=='verified' else 0
        cur.execute("""INSERT INTO certificates (student_id,title,category,issuer,issue_date,file_path,
                       ai_confidence,ai_extracted_name,verification_status,stars_earned,event_title)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (user['id'],title,category,issuer,issue_date,path,ai_conf,user['name'],vstatus,stars,event_title or None))
        mysql.connection.commit()
        if stars: add_stars(user['id'],stars,f'Certificate: {title}','certificate')
        # Notify CC of the student's class
        cc = get_class_coordinator(user.get('class_name',''))
        if cc:
            _notif(cur, cc['id'],
                   f'Certificate Pending: {user["name"]}',
                   f'{title} — {category.replace("_"," ").title()} | Class: {user.get("class_name","")}',
                   'certificate')
            mysql.connection.commit()
        flash(f'Uploaded! Status: {vstatus}','success')
        cur.close()
        return redirect(url_for('student_certificates'))
    cur.execute("SELECT * FROM certificates WHERE student_id=%s ORDER BY created_at DESC", (user['id'],))
    certificates=cur.fetchall(); cur.close()
    return render_template('student/certificates.html', user=user, certificates=certificates)

@app.route('/staff/certificates')
@login_required
@staff_required
def verify_certificates_page():
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()
    # Check if this staff is a CC — if so, show their CC class first
    cur.execute("""SELECT coordinator_class FROM staff_classes
                   WHERE staff_id=%s AND is_coordinator=1 LIMIT 1""", (user['id'],))
    cc_row = cur.fetchone()
    cc_class = cc_row['coordinator_class'] if cc_row else None

    staff_classes = get_staff_classes(user['id'])
    # CC class gets priority; merge with other classes
    all_visible = []
    if cc_class and cc_class not in all_visible:
        all_visible.append(cc_class)
    for c in staff_classes:
        if c not in all_visible:
            all_visible.append(c)

    if not all_visible:
        all_visible = ['__none__']
    ph = ','.join(['%s']*len(all_visible))
    cur.execute(f"""SELECT c.*, u.name AS student_name, u.class_name AS student_class
                   FROM certificates c JOIN users u ON u.id=c.student_id
                   WHERE u.class_name IN ({ph}) ORDER BY
                   FIELD(c.verification_status,'pending','manual_review','verified','rejected'),
                   c.created_at DESC""", tuple(all_visible))
    certificates = cur.fetchall()
    cur.close()
    return render_template('staff/verify_certificates.html',
        user=user, certificates=certificates, cc_class=cc_class)

@app.route('/certificate/<int:cert_id>/<action>')
@login_required
@staff_required
def certificate_action(cert_id, action):
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()
    cur.execute("SELECT * FROM certificates WHERE id=%s", (cert_id,))
    cert = cur.fetchone()
    if cert:
        if action=='approve':
            stars=CERT_STARS.get(cert['category'],5)
            cur.execute("UPDATE certificates SET verification_status='verified',stars_earned=%s,verified_by=%s WHERE id=%s",
                        (stars,user['id'],cert_id))
            mysql.connection.commit()
            add_stars(cert['student_id'],stars,f'Certificate verified: {cert["title"]}','certificate')
            flash('Verified ✓','success')
        elif action=='reject':
            cur.execute("UPDATE certificates SET verification_status='rejected',verified_by=%s WHERE id=%s",
                        (user['id'],cert_id))
            mysql.connection.commit()
            notify(cert['student_id'],'Certificate Rejected',f'"{cert["title"]}" was not approved.','certificate')
            flash('Rejected','warning')
    cur.close()
    return redirect(request.referrer or url_for('verify_certificates_page'))

# ════════════════════════════════════════════
#  PROJECTS, SEMINARS, LIBRARY, RESEARCH
# ════════════════════════════════════════════

@app.route('/student/projects', methods=['GET','POST'])
@login_required
@student_required
def student_projects():
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()
    if request.method == 'POST':
        title=request.form.get('title',''); desc=request.form.get('description','')
        gh=request.form.get('github_link',''); tech=request.form.get('tech_stack','')
        status=request.form.get('status','ongoing'); stars=app.config['STARS']['project_add']
        cur.execute("INSERT INTO projects (student_id,title,description,github_link,tech_stack,status,stars_earned) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                    (user['id'],title,desc,gh,tech,status,stars))
        mysql.connection.commit(); add_stars(user['id'],stars,f'Project: {title}','achievement')
        flash(f'+{stars} ⭐','success'); cur.close(); return redirect(url_for('student_projects'))
    cur.execute("SELECT * FROM projects WHERE student_id=%s ORDER BY created_at DESC",(user['id'],))
    projects=cur.fetchall(); cur.close()
    return render_template('student/projects.html', user=user, projects=projects)

@app.route('/student/seminars', methods=['GET','POST'])
@login_required
@student_required
def student_seminars():
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()
    if request.method == 'POST':
        title=request.form.get('title',''); desc=request.form.get('description','')
        dt=request.form.get('conducted_date') or None; aud=request.form.get('audience','')
        image=request.files.get('image'); img_path=None
        if image and allowed(image.filename,ALLOWED_IMG): img_path=save_file(image,'seminars')
        stars=app.config['STARS']['seminar_add']
        cur.execute("INSERT INTO seminars (student_id,title,description,conducted_date,audience,image_path,stars_earned) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                    (user['id'],title,desc,dt,aud,img_path,stars))
        mysql.connection.commit(); add_stars(user['id'],stars,f'Seminar: {title}','achievement')
        flash(f'+{stars} ⭐','success'); cur.close(); return redirect(url_for('student_seminars'))
    cur.execute("SELECT * FROM seminars WHERE student_id=%s ORDER BY created_at DESC",(user['id'],))
    seminars=cur.fetchall(); cur.close()
    return render_template('student/seminars.html', user=user, seminars=seminars)
@app.route('/staff/research', methods=['GET','POST'])
@login_required
@staff_required
def staff_research():
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()
    if request.method == 'POST':
        title=request.form.get('title',''); jname=request.form.get('journal_name','')
        pub_dt=request.form.get('publication_date') or None; status=request.form.get('status','under_review')
        doi=request.form.get('doi_link','')
        stars=(app.config['STARS']['research_published'] if status=='published' else
               app.config['STARS']['research_presented'] if status=='presented' else 0)
        cur.execute("INSERT INTO research (staff_id,title,journal_name,publication_date,status,doi_link,stars_earned) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                    (user['id'],title,jname,pub_dt,status,doi,stars))
        mysql.connection.commit()
        if stars: add_stars(user['id'],stars,f'Research: {title}','achievement')
        flash('Research added ✓','success'); cur.close(); return redirect(url_for('staff_research'))
    cur.execute("SELECT * FROM research WHERE staff_id=%s ORDER BY created_at DESC",(user['id'],))
    research=cur.fetchall(); cur.close()
    return render_template('staff/research.html', user=user, research=research)

@app.route('/staff/analytics')
@login_required
@staff_required
def staff_analytics():
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()
    staff_classes = get_staff_classes(user['id'])
    active_class  = request.args.get('class') or (staff_classes[0] if staff_classes else '')

    students = []
    att_by_subject = []
    quiz_summary   = []
    assignment_summary = []

    if active_class:
        # Students with per-class attendance (only this staff's records for this class)
        cur.execute("""SELECT u.*,
                       (SELECT COUNT(*) FROM attendance
                        WHERE student_id=u.id AND staff_id=%s AND class_name=%s AND status='present') AS present_count,
                       (SELECT COUNT(*) FROM attendance
                        WHERE student_id=u.id AND staff_id=%s AND class_name=%s) AS total_att,
                       (SELECT COUNT(*) FROM certificates
                        WHERE student_id=u.id AND verification_status='verified') AS certs,
                       (SELECT COUNT(*) FROM projects WHERE student_id=u.id) AS projects_count,
                       (SELECT COUNT(*) FROM submissions s2
                        JOIN assignments a2 ON a2.id=s2.assignment_id
                        WHERE s2.student_id=u.id AND a2.staff_id=%s AND a2.class_name=%s) AS submissions_count
                       FROM users u WHERE u.class_name=%s AND u.role='student'
                       ORDER BY u.streak_stars DESC""",
                    (user['id'], active_class, user['id'], active_class,
                     user['id'], active_class, active_class))
        students = cur.fetchall()

        # Attendance breakdown by subject for this staff + class
        cur.execute("""SELECT subject,
                       SUM(status='present') AS present,
                       SUM(status='absent')  AS absent,
                       SUM(status='late')    AS late,
                       COUNT(*) AS total,
                       ROUND(SUM(status='present')/COUNT(*)*100,1) AS pct
                       FROM attendance
                       WHERE staff_id=%s AND class_name=%s
                       GROUP BY subject ORDER BY subject""",
                    (user['id'], active_class))
        att_by_subject = cur.fetchall()

        # Quiz summary for this staff + class
        cur.execute("""SELECT q.title, q.subject,
                       COUNT(qa.id) AS attempts,
                       ROUND(AVG(qa.percentage),1) AS avg_pct,
                       MAX(qa.percentage) AS max_pct,
                       MIN(qa.percentage) AS min_pct
                       FROM quizzes q
                       LEFT JOIN quiz_attempts qa ON qa.quiz_id=q.id
                       WHERE q.staff_id=%s AND q.class_name=%s
                       GROUP BY q.id ORDER BY q.created_at DESC""",
                    (user['id'], active_class))
        quiz_summary = cur.fetchall()

        # Assignment submission summary
        cur.execute("""SELECT a.title, a.subject, a.due_date,
                       COUNT(s.id) AS submitted,
                       (SELECT COUNT(*) FROM users WHERE class_name=%s AND role='student') AS total_students,
                       ROUND(AVG(s.marks_obtained),1) AS avg_marks
                       FROM assignments a
                       LEFT JOIN submissions s ON s.assignment_id=a.id
                       WHERE a.staff_id=%s AND a.class_name=%s
                       GROUP BY a.id ORDER BY a.created_at DESC LIMIT 10""",
                    (active_class, user['id'], active_class))
        assignment_summary = cur.fetchall()

    low_engagement = [s for s in students if (s['total_att'] or 0) > 0 and
                      (s['present_count'] or 0)/(s['total_att'] or 1)*100 < 75]
    cur.close()
    return render_template('staff/analytics.html',
        user=user, students=students,
        low_engagement=low_engagement, active_class=active_class,
        staff_classes=staff_classes, att_by_subject=att_by_subject,
        quiz_summary=quiz_summary, assignment_summary=assignment_summary)

# ════════════════════════════════════════════
#  STUDENT DASHBOARD + LEADERBOARD
# ════════════════════════════════════════════

def _student_dashboard(user):
    cur = get_cur(); today=date.today()
    cur.execute("SELECT COUNT(*) AS total, SUM(status='present') AS present FROM attendance WHERE student_id=%s",(user['id'],))
    att=cur.fetchone() or {'total':0,'present':0}
    cur.execute("""SELECT a.*, s.id AS submitted, s.status AS sub_status
                   FROM assignments a LEFT JOIN submissions s ON s.assignment_id=a.id AND s.student_id=%s
                   WHERE a.class_name=%s ORDER BY a.due_date""",(user['id'],user['class_name']))
    pending_assignments=cur.fetchall()
    cur.execute("""SELECT q.* FROM quizzes q WHERE q.class_name=%s
                   AND (q.due_date IS NULL OR q.due_date>NOW())
                   AND q.id NOT IN (SELECT quiz_id FROM quiz_attempts WHERE student_id=%s)""",
                (user['class_name'],user['id']))
    active_quizzes=cur.fetchall()
    cur.execute("SELECT id,name,streak_stars FROM users WHERE class_name=%s AND role='student' ORDER BY streak_stars DESC LIMIT 5",(user['class_name'],))
    top5=cur.fetchall()
    cur.execute("SELECT COUNT(*)+1 AS rnk FROM users WHERE class_name=%s AND role='student' AND streak_stars>%s",(user['class_name'],user['streak_stars']))
    rrow=cur.fetchone(); rank=rrow['rnk'] if rrow else None
    cur.execute("SELECT * FROM streak_log WHERE user_id=%s ORDER BY created_at DESC LIMIT 10",(user['id'],))
    recent_activity=cur.fetchall()
    cur.execute("SELECT * FROM notifications WHERE user_id=%s ORDER BY created_at DESC LIMIT 5",(user['id'],))
    notifications=cur.fetchall()
    cur.close()
    level_progress=((user['streak_stars'] or 0)%1000)/10
    return render_template('student/dashboard.html',
        user=user,att=att,today=today,pending_assignments=pending_assignments,
        active_quizzes=active_quizzes,top5=top5,rank=rank,
        recent_activity=recent_activity,notifications=notifications,level_progress=level_progress)

@app.route('/student/profile', methods=['GET','POST'])
@login_required
@student_required
def student_profile():
    user = get_current_user()
    if not user: return redirect(url_for('login'))
    cur = get_cur()

    if request.method == 'POST':
        phone    = request.form.get('phone','')
        pic      = request.files.get('profile_pic')
        pic_path = user.get('profile_pic')
        if pic and allowed(pic.filename, ALLOWED_IMG):
            pic_path = save_file(pic, 'profiles')
        cur.execute("UPDATE users SET phone=%s, profile_pic=%s WHERE id=%s",
                    (phone, pic_path, user['id']))
        mysql.connection.commit()
        flash('Profile updated ✓','success')
        cur.close()
        return redirect(url_for('student_profile'))

    # Stats
    cur.execute("SELECT COUNT(*) AS total, SUM(status='present') AS present FROM attendance WHERE student_id=%s", (user['id'],))
    att = cur.fetchone() or {'total':0,'present':0}
    cur.execute("SELECT COUNT(*) AS c FROM certificates WHERE student_id=%s AND verification_status='verified'", (user['id'],))
    cert_count = cur.fetchone()['c']
    cur.execute("SELECT COUNT(*) AS c FROM projects WHERE student_id=%s", (user['id'],))
    project_count = cur.fetchone()['c']
    cur.execute("SELECT * FROM streak_log WHERE user_id=%s ORDER BY created_at DESC LIMIT 20", (user['id'],))
    streak_history = cur.fetchall()
    cur.execute("SELECT COUNT(*)+1 AS rnk FROM users WHERE class_name=%s AND role='student' AND streak_stars>%s",
                (user['class_name'], user['streak_stars']))
    rank_row = cur.fetchone()
    rank = rank_row['rnk'] if rank_row else '-'
    cur.close()

    att_pct = round((att['present'] or 0)/(att['total'] or 1)*100, 1)
    level_progress = ((user['streak_stars'] or 0) % 1000) / 10

    return render_template('student/profile.html',
        user=user, att=att, att_pct=att_pct,
        cert_count=cert_count, project_count=project_count,
        streak_history=streak_history, rank=rank,
        level_progress=level_progress)

@app.route('/leaderboard')
@login_required
def leaderboard():
    user=get_current_user()
    if not user: return redirect(url_for('login'))
    cur=get_cur()
    cur.execute("SELECT id,name,roll_number,streak_stars,level FROM users WHERE class_name=%s AND role='student' ORDER BY streak_stars DESC",(user['class_name'],))
    all_students=cur.fetchall()
    badges=['Gold','Silver','Bronze']
    for i,s in enumerate(all_students): s['badge']=badges[i] if i<3 else 'Active'
    top5=all_students[:5]; rank=next((i+1 for i,s in enumerate(all_students) if s['id']==user['id']),None)
    cur.close()
    return render_template('student/leaderboard.html',user=user,all_students=all_students,top5=top5,rank=rank)

@app.route('/student/attendance')
@login_required
@student_required
def student_attendance():
    user=get_current_user()
    if not user: return redirect(url_for('login'))
    cur=get_cur()
    cur.execute("""SELECT subject,
                   SUM(status='present') AS present,
                   COUNT(*) AS total
                   FROM attendance WHERE student_id=%s GROUP BY subject ORDER BY subject""",
                (user['id'],))
    subject_summary=cur.fetchall()
    cur.execute("""SELECT a.*, u.name AS staff_name, t.start_time, t.end_time
                   FROM attendance a
                   LEFT JOIN users u ON u.id=a.staff_id
                   LEFT JOIN timetable t ON t.id=a.timetable_id
                   WHERE a.student_id=%s
                   ORDER BY a.date DESC, a.period""",
                (user['id'],))
    records=cur.fetchall()
    cur.close()
    return render_template('student/attendance.html',user=user,subject_summary=subject_summary,records=records)

# ════════════════════════════════════════════
#  NOTIFICATIONS, PROFILE (already above)
# ════════════════════════════════════════════

@app.route('/notifications')
@login_required
def notifications():
    user=get_current_user()
    if not user: return redirect(url_for('login'))
    cur=get_cur()
    cur.execute("SELECT * FROM notifications WHERE user_id=%s ORDER BY created_at DESC LIMIT 50",(user['id'],))
    notifs=cur.fetchall()
    cur.execute("UPDATE notifications SET is_read=1 WHERE user_id=%s",(user['id'],))
    mysql.connection.commit(); cur.close()
    return render_template('notifications.html',user=user,notifications=notifs)

@app.route('/api/notifications/count')
@login_required
def notif_count_api():
    return jsonify({'count':get_notif_count(session['user_id'])})

@app.route('/resume')
@login_required
@student_required
def generate_resume_page():
    user=get_current_user()
    if not user: return redirect(url_for('login'))
    cur=get_cur()
    cur.execute("SELECT * FROM certificates WHERE student_id=%s AND verification_status='verified'",(user['id'],)); certs=cur.fetchall()
    cur.execute("SELECT * FROM projects WHERE student_id=%s",(user['id'],)); projects=cur.fetchall()
    cur.execute("SELECT * FROM seminars WHERE student_id=%s",(user['id'],)); seminars=cur.fetchall()
    cur.execute("SELECT * FROM semester_results WHERE student_id=%s ORDER BY semester, subject",(user['id'],)); results=cur.fetchall()
    cur.close()
    try:
        from fpdf import FPDF
        pdf=FPDF(); pdf.add_page(); pdf.set_font('Helvetica','B',20)
        pdf.cell(0,10,user['name'],ln=True); pdf.set_font('Helvetica','',12)
        pdf.cell(0,7,f"{user['department']} | {user['class_name'] or ''} | {user['roll_number'] or ''}",ln=True)
        pdf.cell(0,7,f"Email: {user['email']}  Stars: {user['streak_stars']} | Level: {user['level']}",ln=True)
        def section(t):
            pdf.set_font('Helvetica','B',13); pdf.set_fill_color(220,220,220)
            pdf.cell(0,8,t,ln=True,fill=True); pdf.set_font('Helvetica','',11)
        if certs:
            section('Certificates')
            for c in certs:
                line = f"- {c['title']} ({c['category'].replace('_',' ').title()})"
                if c.get('event_title'):
                    line += f" | {c['event_title']}"
                pdf.cell(0,6,line,ln=True)
        if projects: section('Projects'); [pdf.cell(0,6,f"- {p['title']} [{p['tech_stack'] or ''}]",ln=True) for p in projects]
        if seminars: section('Seminars'); [pdf.cell(0,6,f"- {s['title']}",ln=True) for s in seminars]
        if results: section('Results'); [pdf.cell(0,6,f"- Sem{r['semester']} {r['subject']} ({r.get('subject_type','theory').title()}) — Grade: {r['grade']}",ln=True) for r in results]
        path=os.path.join(TEMP_DIR, f'resume_{user["id"]}.pdf'); pdf.output(path)
        return send_file(path,as_attachment=True,download_name=f"{user['name'].replace(' ','_')}_Resume.pdf",mimetype='application/pdf')
    except ImportError:
        flash('pip install fpdf2','warning'); return redirect(url_for('profile'))




if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
